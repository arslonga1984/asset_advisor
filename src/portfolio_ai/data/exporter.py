"""Excel export for analysis results."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from portfolio_ai.core.models import AnalysisResult


def export_to_excel(result: AnalysisResult, path: Path) -> None:
    """Export analysis result to a multi-sheet Excel file."""
    wb = openpyxl.Workbook()

    _write_summary_sheet(wb, result)
    _write_holdings_sheet(wb, result)
    _write_metrics_sheet(wb, result)

    if result.ai_insights:
        _write_insights_sheet(wb, result)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(str(path))


def _write_summary_sheet(wb: openpyxl.Workbook, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Summary")
    rows = [
        ("Portfolio Name", result.portfolio_name),
        ("Currency", result.currency),
        ("Benchmark", result.benchmark_ticker),
        ("Total Value", result.total_value),
        ("Total Cost", result.total_cost),
        ("Profit/Loss", result.profit_loss),
        ("Total Return (%)", result.total_return_pct),
        ("Annualized Return (%)", result.annualized_return_pct),
    ]
    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)


def _write_holdings_sheet(wb: openpyxl.Workbook, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Holdings")
    headers = [
        "Ticker", "Name", "Quantity", "Avg Cost", "Current Price",
        "Market Value", "Weight (%)", "Return (%)", "Sector",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, h in enumerate(result.holdings_analysis, start=2):
        ws.cell(row=row_idx, column=1, value=h.ticker)
        ws.cell(row=row_idx, column=2, value=h.name)
        ws.cell(row=row_idx, column=3, value=h.quantity)
        ws.cell(row=row_idx, column=4, value=h.avg_cost)
        ws.cell(row=row_idx, column=5, value=h.current_price)
        ws.cell(row=row_idx, column=6, value=h.market_value)
        ws.cell(row=row_idx, column=7, value=h.weight_pct)
        ws.cell(row=row_idx, column=8, value=h.return_pct)
        ws.cell(row=row_idx, column=9, value=h.sector)


def _write_metrics_sheet(wb: openpyxl.Workbook, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Metrics")
    metrics = [
        ("Volatility (%)", result.volatility_pct),
        ("Sharpe Ratio", result.sharpe_ratio),
        ("Max Drawdown (%)", result.max_drawdown_pct),
        ("Beta", result.beta),
        ("Alpha (%)", result.alpha_pct),
    ]
    ws.cell(row=1, column=1, value="Metric")
    ws.cell(row=1, column=2, value="Value")
    for row_idx, (label, value) in enumerate(metrics, start=2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)


def _write_insights_sheet(wb: openpyxl.Workbook, result: AnalysisResult) -> None:
    ws = wb.create_sheet("Insights")
    ws.cell(row=1, column=1, value="AI Insights")
    ws.cell(row=2, column=1, value=result.ai_insights)
