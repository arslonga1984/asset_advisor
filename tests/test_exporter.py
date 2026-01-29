"""Tests for Excel exporter."""

from pathlib import Path

import openpyxl
import pytest

from portfolio_ai.core.models import AnalysisResult, HoldingAnalysis
from portfolio_ai.data.exporter import export_to_excel


@pytest.fixture
def sample_result():
    holdings = (
        HoldingAnalysis(
            ticker="AAPL",
            name="Apple Inc.",
            quantity=10,
            avg_cost=150.0,
            current_price=175.0,
            market_value=1750.0,
            weight_pct=50.0,
            return_pct=16.67,
            sector="Technology",
        ),
        HoldingAnalysis(
            ticker="MSFT",
            name="Microsoft Corp.",
            quantity=5,
            avg_cost=300.0,
            current_price=350.0,
            market_value=1750.0,
            weight_pct=50.0,
            return_pct=16.67,
            sector="Technology",
        ),
    )
    return AnalysisResult(
        portfolio_name="Test Portfolio",
        total_value=3500.0,
        total_cost=3000.0,
        total_return_pct=16.67,
        annualized_return_pct=16.67,
        volatility_pct=20.0,
        sharpe_ratio=0.73,
        max_drawdown_pct=-8.0,
        beta=1.1,
        alpha_pct=2.0,
        currency="USD",
        holdings_analysis=holdings,
        benchmark_ticker="SPY",
        ai_insights="Test insight.",
    )


class TestExportToExcel:
    def test_creates_file(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        assert output.exists()

    def test_has_summary_sheet(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        wb = openpyxl.load_workbook(output)
        assert "Summary" in wb.sheetnames

    def test_has_holdings_sheet(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        wb = openpyxl.load_workbook(output)
        assert "Holdings" in wb.sheetnames

    def test_has_metrics_sheet(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        wb = openpyxl.load_workbook(output)
        assert "Metrics" in wb.sheetnames

    def test_has_insights_sheet(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        wb = openpyxl.load_workbook(output)
        assert "Insights" in wb.sheetnames

    def test_holdings_data_correct(self, sample_result, tmp_path: Path):
        output = tmp_path / "output.xlsx"
        export_to_excel(sample_result, output)
        wb = openpyxl.load_workbook(output)
        ws = wb["Holdings"]
        assert ws.cell(row=2, column=1).value == "AAPL"
        assert ws.cell(row=3, column=1).value == "MSFT"

    def test_no_insights_sheet_when_none(self, tmp_path: Path):
        result = AnalysisResult(
            portfolio_name="Test",
            total_value=1000.0,
            total_cost=800.0,
            total_return_pct=25.0,
            annualized_return_pct=25.0,
            volatility_pct=15.0,
            sharpe_ratio=1.0,
            max_drawdown_pct=-5.0,
            beta=1.0,
            alpha_pct=3.0,
            currency="USD",
            holdings_analysis=(),
            benchmark_ticker="SPY",
        )
        output = tmp_path / "output.xlsx"
        export_to_excel(result, output)
        wb = openpyxl.load_workbook(output)
        assert "Insights" not in wb.sheetnames
